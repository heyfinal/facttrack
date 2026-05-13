"""Excel workbook rendering — the landman's exportable artifact.

Tabs:
  1. Cover (project meta + summary stats)
  2. Curative — ranked open items, severity color-coded
  3. Leases — every lease in scope with term + clauses
  4. Chain Events — runback per lease
  5. Wells — production wells + status + operator history
  6. Lease Calendar — 24-month expiration / risk view
"""
from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from facttrack.config import PATHS, ensure_dirs
from facttrack.db import cursor
from facttrack.engine.context import load_project

log = logging.getLogger(__name__)


_HEADER_FILL = PatternFill(start_color="0F3057", end_color="0F3057", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_BODY_WRAP = Alignment(vertical="top", wrap_text=True)

_SEV_FILLS = {
    "critical": PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid"),
    "high":     PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),
    "medium":   PatternFill(start_color="F9A825", end_color="F9A825", fill_type="solid"),
    "low":      PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid"),
}
_SEV_FONT_WHITE = Font(bold=True, color="FFFFFF")
_SEV_FONT_DARK = Font(bold=True, color="222222")


def _header_row(ws, headers: list[str], row: int = 1) -> None:
    for col_idx, label in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[row].height = 22


def _autosize(ws, max_widths: list[int]) -> None:
    for idx, w in enumerate(max_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _fmt(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def render_xlsx(project_id: str) -> Path:
    ensure_dirs()
    ctx = load_project(project_id)
    wb = Workbook()

    # ── Cover ────────────────────────────────────────────────────────
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "FactTrack — East Texas Landwork Report"
    cover["A1"].font = Font(size=18, bold=True, color="0F3057")
    cover["A2"] = ctx.label
    cover["A2"].font = Font(size=12, italic=True, color="555555")
    cover["A4"] = "Project ID"
    cover["B4"] = ctx.project_id
    cover["A5"] = "Generated"
    cover["B5"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    cover["A6"] = "Tracts"
    cover["B6"] = len(ctx.tracts)
    cover["A7"] = "Leases"
    cover["B7"] = len(ctx.leases)
    cover["A8"] = "Wells"
    cover["B8"] = len(ctx.wells)
    cover.column_dimensions["A"].width = 18
    cover.column_dimensions["B"].width = 60
    for r in range(4, 9):
        cover[f"A{r}"].font = Font(bold=True)

    # ── Curative ─────────────────────────────────────────────────────
    cur_ws = wb.create_sheet("Curative")
    headers = ["#", "Severity", "Rule", "Title", "Description", "Suggested Action",
               "Assignee", "$ Low", "$ High", "Confidence", "Deadline", "Status"]
    _header_row(cur_ws, headers)
    cur_ws.freeze_panes = "A2"
    cur_ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    with cursor() as cur:
        cur.execute(
            """
            SELECT rule_id, severity, confidence_score, title, description, suggested_action,
                   assignee_level, dollar_impact_low, dollar_impact_high, deadline, status
            FROM curative_item
            WHERE project_id = %s
            ORDER BY
              CASE severity WHEN 'critical' THEN 4 WHEN 'high' THEN 3
                            WHEN 'medium' THEN 2 WHEN 'low' THEN 1 ELSE 0 END DESC,
              confidence_score DESC,
              dollar_impact_high DESC NULLS LAST
            """,
            (project_id,),
        )
        rows = cur.fetchall()

    for idx, r in enumerate(rows, 1):
        row_idx = idx + 1
        cur_ws.cell(row=row_idx, column=1, value=idx).alignment = Alignment(horizontal="center")
        sev_cell = cur_ws.cell(row=row_idx, column=2, value=r["severity"].upper())
        sev_cell.fill = _SEV_FILLS.get(r["severity"], _SEV_FILLS["low"])
        sev_cell.font = _SEV_FONT_DARK if r["severity"] == "medium" else _SEV_FONT_WHITE
        sev_cell.alignment = Alignment(horizontal="center", vertical="center")
        cur_ws.cell(row=row_idx, column=3, value=r["rule_id"]).alignment = _BODY_WRAP
        cur_ws.cell(row=row_idx, column=4, value=r["title"]).alignment = _BODY_WRAP
        cur_ws.cell(row=row_idx, column=5, value=r["description"]).alignment = _BODY_WRAP
        cur_ws.cell(row=row_idx, column=6, value=r["suggested_action"]).alignment = _BODY_WRAP
        cur_ws.cell(row=row_idx, column=7, value=_fmt(r["assignee_level"])).alignment = _BODY_WRAP
        cur_ws.cell(row=row_idx, column=8, value=float(r["dollar_impact_low"]) if r["dollar_impact_low"] is not None else None)
        cur_ws.cell(row=row_idx, column=8).number_format = '"$"#,##0'
        cur_ws.cell(row=row_idx, column=9, value=float(r["dollar_impact_high"]) if r["dollar_impact_high"] is not None else None)
        cur_ws.cell(row=row_idx, column=9).number_format = '"$"#,##0'
        cur_ws.cell(row=row_idx, column=10, value=float(r["confidence_score"]))
        cur_ws.cell(row=row_idx, column=10).number_format = "0.00"
        cur_ws.cell(row=row_idx, column=11, value=_fmt(r["deadline"]))
        cur_ws.cell(row=row_idx, column=12, value=r["status"])

        cur_ws.row_dimensions[row_idx].height = 60

    _autosize(cur_ws, [5, 11, 28, 40, 70, 70, 16, 12, 12, 11, 12, 10])

    # ── Leases ───────────────────────────────────────────────────────
    le_ws = wb.create_sheet("Leases")
    _header_row(le_ws, [
        "Instrument", "Recording date", "Lessor", "Lessee",
        "Effective date", "Primary term (yr)", "Term end",
        "Royalty fraction", "Pugh", "Retained acreage", "Cont dev",
        "Depth limit (ft)", "Confidence",
    ])
    le_ws.freeze_panes = "A2"
    for i, le in enumerate(ctx.leases, 1):
        le_ws.cell(row=i + 1, column=1, value=le.opr_instrument_no)
        le_ws.cell(row=i + 1, column=2, value=_fmt(le.recording_date))
        le_ws.cell(row=i + 1, column=3, value=le.lessor_text)
        le_ws.cell(row=i + 1, column=4, value=le.lessee_text)
        le_ws.cell(row=i + 1, column=5, value=_fmt(le.effective_date))
        le_ws.cell(row=i + 1, column=6, value=float(le.primary_term_years) if le.primary_term_years else None)
        le_ws.cell(row=i + 1, column=7, value=_fmt(le.primary_term_end))
        le_ws.cell(row=i + 1, column=8, value=float(le.royalty_fraction) if le.royalty_fraction else None)
        le_ws.cell(row=i + 1, column=8).number_format = "0.0000"
        le_ws.cell(row=i + 1, column=9, value="Yes" if le.has_pugh_clause else "No")
        le_ws.cell(row=i + 1, column=10, value="Yes" if le.has_retained_acreage else "No")
        le_ws.cell(row=i + 1, column=11, value="Yes" if le.has_continuous_dev else "No")
        le_ws.cell(row=i + 1, column=12, value=float(le.depth_limit_ft) if le.depth_limit_ft else None)
        le_ws.cell(row=i + 1, column=13, value=float(le.confidence_score) if le.confidence_score else None)
    _autosize(le_ws, [16, 14, 30, 30, 14, 9, 14, 10, 8, 10, 10, 12, 10])

    # ── Chain Events ────────────────────────────────────────────────
    ev_ws = wb.create_sheet("Chain Events")
    _header_row(ev_ws, [
        "Instrument", "Recording date", "Type", "Grantor", "Grantee", "Lease ref", "Confidence",
    ])
    ev_ws.freeze_panes = "A2"
    for i, ev in enumerate(ctx.chain_events, 1):
        ev_ws.cell(row=i + 1, column=1, value=ev.opr_instrument_no)
        ev_ws.cell(row=i + 1, column=2, value=_fmt(ev.recording_date))
        ev_ws.cell(row=i + 1, column=3, value=ev.event_type)
        ev_ws.cell(row=i + 1, column=4, value=ev.grantor_text)
        ev_ws.cell(row=i + 1, column=5, value=ev.grantee_text)
        ev_ws.cell(row=i + 1, column=6, value=ev.references_lease_id)
    _autosize(ev_ws, [16, 14, 16, 30, 30, 10, 10])

    # ── Wells ───────────────────────────────────────────────────────
    well_ws = wb.create_sheet("Wells")
    _header_row(well_ws, [
        "API #", "Lease name", "Well #", "Operator P5", "Operator name",
        "Status", "Spud date", "Completion date",
    ])
    well_ws.freeze_panes = "A2"
    for i, w in enumerate(ctx.wells, 1):
        op_name = ctx.operators_by_p5.get(w.operator_p5).name if (w.operator_p5 and w.operator_p5 in ctx.operators_by_p5) else ""
        well_ws.cell(row=i + 1, column=1, value=w.api_no)
        well_ws.cell(row=i + 1, column=2, value=w.lease_name)
        well_ws.cell(row=i + 1, column=3, value=w.well_no)
        well_ws.cell(row=i + 1, column=4, value=w.operator_p5)
        well_ws.cell(row=i + 1, column=5, value=op_name)
        well_ws.cell(row=i + 1, column=6, value=w.status)
        well_ws.cell(row=i + 1, column=7, value=_fmt(w.spud_date))
        well_ws.cell(row=i + 1, column=8, value=_fmt(w.completion_date))
    _autosize(well_ws, [18, 22, 8, 12, 30, 12, 14, 14])

    # ── Lease Calendar ───────────────────────────────────────────────
    cal_ws = wb.create_sheet("Lease Calendar")
    _header_row(cal_ws, [
        "Instrument", "Lessor", "Effective", "Primary term end",
        "Days remaining", "Risk", "Pugh", "Continuous dev",
    ])
    cal_ws.freeze_panes = "A2"
    today = date.today()
    for i, le in enumerate(ctx.leases, 1):
        cal_ws.cell(row=i + 1, column=1, value=le.opr_instrument_no)
        cal_ws.cell(row=i + 1, column=2, value=le.lessor_text)
        cal_ws.cell(row=i + 1, column=3, value=_fmt(le.effective_date))
        cal_ws.cell(row=i + 1, column=4, value=_fmt(le.primary_term_end))
        if le.primary_term_end:
            delta = (le.primary_term_end - today).days
            cal_ws.cell(row=i + 1, column=5, value=delta)
            risk_color = "green" if delta > 180 else ("yellow" if delta > 0 else "red")
            sev_label = {"green": "Active", "yellow": "Approaching", "red": "Past term"}[risk_color]
            risk_cell = cal_ws.cell(row=i + 1, column=6, value=sev_label)
            risk_cell.fill = {
                "green": PatternFill("solid", "C8E6C9"),
                "yellow": PatternFill("solid", "FFF59D"),
                "red": PatternFill("solid", "FFCDD2"),
            }[risk_color]
        cal_ws.cell(row=i + 1, column=7, value="Yes" if le.has_pugh_clause else "No")
        cal_ws.cell(row=i + 1, column=8, value="Yes" if le.has_continuous_dev else "No")
    _autosize(cal_ws, [16, 30, 14, 16, 14, 14, 8, 14])

    out_dir = PATHS.reports / project_id
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / "report.xlsx"
    wb.save(xlsx_path)
    log.info("rendered Excel → %s", xlsx_path)
    return xlsx_path


if __name__ == "__main__":  # pragma: no cover
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--project", required=True)
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    path = render_xlsx(args.project)
    print(f"Excel: {path}")
